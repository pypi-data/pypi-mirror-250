# -*- coding: utf-8 -*-
"""
This module offers a set of operations that user can modify their excel files.
"""
import pathlib
import warnings
from os import remove
from pathlib import Path
from typing import List, Optional, Tuple, Union

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from urllib3 import disable_warnings

from .excel_definition import ExcelDefinition
from .sprint_schedule import SprintScheduleStore
from .story import Story, StoryFactory
from .utils import is_absolute_path_valid

__all__ = ["read_excel_file", "output_to_excel_file"]

# Currently, the openpyxl package will report an obsolete warning.
warnings.simplefilter(action="ignore", category=UserWarning)
# Disable the HTTPS certificate verification warning.
disable_warnings()


def read_excel_file(
    file: Union[str, Path],
    excel_definition: ExcelDefinition,
    sprint_schedule: SprintScheduleStore,
) -> Tuple[List[str], List[Story]]:
    """
    Read and parse the excel file

    :parm file:
        The excel file that you want to read

    :parm excel_definition:
        The excel column definition which is imported
        from the :py:class:`ExcelDefinition`

    :parm sprint_schedule:
        The priority mapping for the :py:class:`Milestone` object.

    :return:
        A :py:class:`tuple` object which contains a list of column
        name and a list of :py:class:`Story`.
    """
    if not is_absolute_path_valid(file):
        raise FileNotFoundError(
            f"""Please make sure the input excel file exist and 
            the path should be absolute. File: {file}."""
        )

    with open(str(file), mode="rb") as raw_file:
        work_book: Workbook = openpyxl.load_workbook(
            raw_file, read_only=True, keep_vba=False, data_only=True, keep_links=True
        )

        if work_book.active is None or (
            not isinstance(work_book.active, Worksheet)
            and not isinstance(work_book.active, ReadOnlyWorksheet)
        ):
            work_book.close()
            raise ValueError("The input excel file doesn't contain any sheets.")

        sheet: Union[Worksheet, ReadOnlyWorksheet] = work_book.active

        # The count of column is taking from the definition file to avoid too
        # many columns inside the excel file. Also, need to avoid exceed
        # the range of the actual count.
        column_count = excel_definition.max_column_index
        if sheet.max_column is not None:
            column_count = min(excel_definition.max_column_index, sheet.max_column)

        if sheet.max_row is not None and sheet.max_row < 2:
            work_book.close()
            return ([], [])

        columns: List[str] = []

        for column_index in range(1, column_count + 1):
            columns.append(str(sheet.cell(row=1, column=column_index).value))

        stories = []

        excel_definition_columns = excel_definition.get_columns()
        story_factory = StoryFactory(excel_definition_columns)

        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(columns)
        ):
            if _should_skip(row):
                continue

            story: Story = story_factory.create_story()
            for column_index, column in enumerate(row):
                definition_column = excel_definition_columns[column_index]
                if definition_column["name"] is None:
                    continue
                story.set_value(
                    definition_column["type"], definition_column["name"], column.value
                )

            story.calc_sprint_schedule(sprint_schedule)
            stories.append(story)

        work_book.close()
        raw_file.close()
    return (columns, stories)


def _should_skip(row: tuple) -> bool:
    if len(row) == 0:
        return True
    first_cell_value = row[0].value
    if first_cell_value is None or len(str(first_cell_value)) == 0:
        return True
    return False


def output_to_excel_file(
    file: Union[str, Path],
    stories: "List[Story]",
    excel_definition: ExcelDefinition,
    columns_in_excel: Optional[List[str]] = None,
    over_write: bool = True,
):
    """
    Generate excel file

    :parm file:
        Output excel file name including the path

    :parm stories:
        A list of :py:class:`Story` which need to be wrote to the excel

    :parm excel_definition:
        The excel column definition which is imported from
        the :py:class:`ExcelDefinition`

    :parm columns_in_excel:
        Using separate column names instead of importing from
        the :py:class:`ExcelDefinition`. Usually, it comes from the
        input excel file.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError("The output file path is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            try:
                remove(file)
            except PermissionError as e:
                raise FileExistsError(
                    f"The exist excel file: {file} cannot be removed. {e.args[0]}"
                ) from e
        else:
            raise FileExistsError(f"The output excel file: {file} is already exist.")

    work_book = openpyxl.Workbook(write_only=False)

    if work_book.active is None or (
        not isinstance(work_book.active, Worksheet)
        and not isinstance(work_book.active, Worksheet)
    ):
        work_book.close()
        raise ValueError("The output excel file cannot be generated.")

    sheet: Worksheet = work_book.active

    excel_definition_columns = excel_definition.get_columns()

    # Use original excel column name first.
    columns = columns_in_excel
    if columns is None:
        columns = [column["name"] for column in excel_definition_columns]

    for column_index, column in enumerate(columns):
        cell = sheet.cell(row=1, column=column_index + 1)
        # There are three kinds of Cells. Only the Cell has the value attribute.
        if hasattr(cell, "value"):
            setattr(cell, "value", column)

    if stories is not None and stories:
        for row_index, story in enumerate(stories):
            for column in excel_definition_columns:
                if column["name"] is None:
                    continue
                cell = sheet.cell(row=row_index + 2, column=column["index"])
                if hasattr(cell, "value"):
                    setattr(cell, "value", story.format_value(column["name"]))

    work_book.save(str(file))
    work_book.close()
