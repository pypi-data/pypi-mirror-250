from tkinter import E
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


VERBOSE = False

# Vacía todas las celdas de una hoja en excel dejando la cabecera de la fila 1
def limpia_hoja_excel(hoja):
    for fila in hoja.iter_rows(min_row=2):
        for celda in fila:
            celda.value = ""

# Inserta una nueva fila en la posición de la fila 2 y la rellena en el primer campo con la clave y en el resto con el resto de campos
# Para aprovechamiento retorna retorno para su uso anidado en otras funciones sobre todo para valores de clave como numérico y str
def escribe_rango_excel(hoja, clave, resto, retorno):
    hoja.insert_rows(2)
    ind = 1
    _ = hoja.cell(column=ind, row=2, value="{0}".format(clave))
    ind += 1
    for col, value in resto.items():
        _ = hoja.cell(column=ind, row=2, value="{0}".format(value))
        ind += 1
    return(retorno)

# Lee de una hoja excel y un archivo excel y la carga en un diccionario usando como clave la concatenación de la lista de campos_clave
def leer_excel(nombre_archivo, nombre_hoja , campos_clave: list):
    book = load_workbook(filename=nombre_archivo, read_only=True)
    hoja = book[nombre_hoja]
    # print('dimensiones:', hoja.calculate_dimension())
    titulos = hoja[1]
    columnas_clave = []
    dict = {}
    for titulo in titulos:
        for campo in campos_clave:
            if titulo.value == campo:
                # print('campo encontrado:', titulo.column_letter)
                columnas_clave.append(titulo.column_letter)

    # print(columnas_clave)
    if len(columnas_clave) != len(campos_clave):
        print('ERROR: No se ha podido encontrar alguno de los campos clave en el Excel.\nSe interrumpe la lectura.')
        return

    # dict['titulos'] = [i.value for i in titulos]

    num_claves_repetidas = 0
    for fila in hoja.iter_rows(min_row=2):
        indice_titulos = 0
        clave = ''
        dict_valores = {}
        for celda in fila:
            # En el caso de que haya celdas sin valor de columna
            if(celda.value is None):  # celda.data_type == 'n'
                #print('AAAAAAAAA celda vacía:', celda.value, fila)
                dict_valores[titulos[indice_titulos].value] = ''
                indice_titulos += 1
                continue
            elif celda.column_letter in columnas_clave:
                # print('celda clave', celda.value)
                if isinstance(celda.value, datetime.datetime):
                    clave = clave + \
                        celda.value.strftime('%d/%m/%Y %H:%M:%S') + ';'
                else:
                    clave = clave + str(celda.value) + ';'

            dict_valores[titulos[indice_titulos].value] = celda.value
            indice_titulos += 1
            # print(celda.coordinate)
        clave = clave[:-1]
        #print('clave:', clave)
        if clave not in dict:
            dict[clave] = dict_valores
        else:
            if VERBOSE:
                print('AVISO: clave ya existente en el diccionario, no se añade la fila:',
                      clave, '-', dict_valores)
            num_claves_repetidas += 1
    book.close()

    lista_vacios = []
    for clave in dict:
        if clave == '':
            lista_vacios.append(clave)

    for item in lista_vacios:
        dict.pop(item)
    # Se calcula el numero total de filas leídas sumando el numero de claves en el diccionario,
    # el numero de filas no añadidas debido a repetición de clave
    # y una más que representa la fila de títulos
    if VERBOSE:
        if (list(dict.values())):
            print(f'\nSe ha leído la hoja "{nombre_hoja}" en el fichero excel "{nombre_archivo}", usando como claves: {campos_clave}\n' +
                f'Se han leído {len(dict.keys()) + num_claves_repetidas + 1} filas y {len(list(dict.values())[0])} columnas\n' +
                f'Hay {len(dict.keys())} claves únicas, y se omitieron {num_claves_repetidas} filas debido a repetición de clave\n')
    return dict




def escribir_excel(rutaExcel, nombreHoja, dict_final):
    # Cargas el Excel sino lo creas
    try:
        book = load_workbook(filename=rutaExcel)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = nombreHoja
        wb.save(filename = rutaExcel)
        book = load_workbook(filename=rutaExcel)
        print(f'Se crea el archivo de facturas {rutaExcel}')

    # Activamos la hoja a usar
    hoja = book[nombreHoja]
    # Eliminamos primera fila por si existiesen titulos
    hoja.delete_rows(1, hoja.max_row)

    # Añadimos titulos
    titulos = ""
    for tit in dict_final.values():
        titulos = tit.keys()
    hoja.append(list(titulos))
            
    # Añadimos filas
    for item in dict_final.values():
        hoja.append(list(item.values()))
    
    # Guardamos
    book.save(filename=rutaExcel)
    book.close()
    if VERBOSE:
        print(f"Guardado diccionario en la siguiente hoja: {hoja} del Excel: {rutaExcel}")
