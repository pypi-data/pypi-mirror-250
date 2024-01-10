from openpyxl import load_workbook
import sys
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import win32com.client as win32
from pathlib import Path
import os
from copy import copy

def convierte_xls_to_xlsx(fichero):

    if Path(fichero+"m").exists():
        os.remove(fichero+"m")
    if Path(fichero+"x").exists():
        os.remove(fichero+"x")

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(fichero)

    #wb.SaveAs(fichero+"x", FileFormat = 51, )    #FileFormat = 51 is for .xlsx extension
    wb.SaveAs(fichero+"m", FileFormat = 52, )     #FileFormat = 52 is for .xlsm extension with macros
    wb.Close()                                    #FileFormat = 56 is for .xls extension
    excel.Application.Quit()


def unifica_hojas(archivo, desplazamiento):
    print("Conviertiendo de xls a xlsx")
    convierte_xls_to_xlsx(archivo)

    print("Unificando hojas en una sóla hoja")
    book = load_workbook(filename=archivo+"m")
    hojas = book.sheetnames
    print(hojas)
    destino = book[hojas[0]]
    if len(hojas) > 1:   # Hay mas de una hoja
        for org in hojas[1:]:
            origen = book[org]
            if org.lower().find("macro") < 0:
                print("De ", org, " a ", hojas[0], " en ", destino.max_row + 1)
                fila_destino = destino.max_row
                for fila in origen.iter_rows(min_row = desplazamiento):
                    fila_destino += 1
                    for cell in fila:
                        destino.cell(row = fila_destino, column = cell.column).value = cell.value
                        destino.cell(row = fila_destino, column = cell.column).number_format = copy(cell.number_format)     
            book.remove(origen)

    if desplazamiento == 4:  # Hay cabecera de búsqueda
        print("Eliminando cabecera de búsqueda")
        destino.delete_rows(1, amount=2)
    book.save(archivo+"x")
    if Path(archivo+"m").exists():
        os.remove(archivo+"m")


'''
if len(sys.argv) != 3:
    print("Error: usage Convert_XLS_to_XLSX file.xls eliminar_busqueda|no_eliminar_busqueda")
    unifica_hojas("c:\\local\\Docuconta_2022.xls", 4)
else:
    fichero = sys.arv[1]
    busqueda = sys.argv[2]
    if busqueda == "eliminar_busqueda":
        unifica_hojas(fichero, 4)
    else:
        unifica_hojas(fichero, 2)
'''
