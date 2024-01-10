import openpyxl
from openpyxl.utils import get_column_letter
from tsjCLI import mkfile

class Excel:
    
    def __init__(self, save_path, title_line):
        # example.xlsx
        self.save_path = save_path     
        mkfile(save_path)
        # Create a workbook with case sensitivity
        wb = openpyxl.Workbook()
        # Get current active sheet
        ws = wb.active 

        # Change title 
        ws.title = 'summary'     
        self.wb = wb 
        
        self.append("summary", title_line)
        
        # wb.save(self.save_path)  
        pass
    
    def set_column_width(self, ws, width=1, count=10):
        # Set width of columns A to J
        for col in range(1, count):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width * 28.3464567

    def create_case_sensitive_sheet(self, wb, sheet_name):
        existing_names = [name.lower() for name in wb.sheetnames]
        new_name = sheet_name
        count = 1
        while new_name.lower() in existing_names:
            new_name = f"{sheet_name}{count}"
            count += 1
        wb.create_sheet(new_name)
        return new_name
    
    def append(self, sheet_name, add_line):
        # Load the workbook
        # wb = openpyxl.load_workbook(self.save_path) 
        wb = self.wb
        # Check if sheet_name exists
        sheet_name = sheet_name.lower()
        if sheet_name not in wb.sheetnames:
            # Create Sheet1
            # wb.create_sheet(sheet_name)
            sheet_name = self.create_case_sensitive_sheet(wb,sheet_name)
        # print(wb.sheetnames)
        # Attention: openpyxl's names are not case sensitive.    
        # code:https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/workbook/workbook.html
        to_sheet = wb.get_sheet_by_name(sheet_name)
        # Get the active worksheet
        wb.active = to_sheet
        ws = wb.active
        self.set_column_width(ws,1, len(add_line))
        # Append two new rows
        # ws.append(['Value1', 'Value2'])
        ws.append(add_line)
    
    def save(self):
        self.wb.save(self.save_path)
        